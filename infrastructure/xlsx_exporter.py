from openpyxl import Workbook
from datetime import datetime

async def save_to_excel(data, sheet_name, field_mapping):
    if not data:
        print("No data to export.")
        return

    date_str = datetime.now().strftime("%d%m%Y")
    file_name = f"{sheet_name}_{date_str}.xlsx"
    print(f"Saving to {file_name}")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    for col_num, header in enumerate(field_mapping.values(), start=1):
        sheet.cell(row=1, column=col_num, value=header)
        
    for row_num, entry in enumerate(data, start=2):
        for col_num, (field, header) in enumerate(field_mapping.items(), start=1):
            value = entry.get(field, "")
            sheet.cell(row=row_num, column=col_num, value=value)

    workbook.save(file_name)
    print(f"Dados salvos no arquivo: {file_name}")
